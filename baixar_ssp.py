"""
baixar_ssp.py — versão GitHub Actions
=====================================
Baixa os arquivos de veículos subtraídos da SSP-SP (2017 até o ano atual),
normaliza as colunas e salva em data/Veículos Roubados/.

ESTRATÉGIA INTELIGENTE:
  - Anos antigos já salvos: PULA (dados históricos não mudam)
  - Ano atual: sempre re-baixa (pegar meses novos)
  - Novos anos (ex: 2027): baixa automaticamente quando aparecerem

Uso:
    python baixar_ssp.py          # baixa só o que falta
    python baixar_ssp.py --force  # força re-download do ano atual
"""

import requests
import pandas as pd
import openpyxl
import os
import time
import argparse
from datetime import datetime


PASTA_DESTINO = os.path.join("data", "Veículos Roubados")
PASTA_TEMP    = os.path.join("data", "temp_ssp")
BASE_URL      = "https://www.ssp.sp.gov.br/assets/estatistica/transparencia/baseDados/veiculosSub"
HEADERS       = {
    "User-Agent": "Mozilla/5.0",
    "Referer"   : "https://www.ssp.sp.gov.br/estatistica/consultas",
}
ANO_ATUAL = datetime.now().year
ANOS      = list(range(2017, ANO_ATUAL + 1))

REMOVER  = {"VERSAO", "LOGRADOURO_VERSAO", "DESC_NATUREZA_LOCAL"}
RENOMEAR = {
    "MES_REGISTRO_BO"  : "MES",
    "ANO_REGISTRO_BO"  : "ANO",
    "DESCR_COR_VEICULO": "DESC_COR_VEICULO",
}


def ja_existe_normalizado(ano):
    """Verifica se o XLSX normalizado já está no destino."""
    destino = os.path.join(PASTA_DESTINO, f"VeiculosSubtraidos_{ano}.xlsx")
    return os.path.exists(destino)


def baixar(ano):
    """Baixa o XLSX bruto da SSP."""
    os.makedirs(PASTA_TEMP, exist_ok=True)
    destino = os.path.join(PASTA_TEMP, f"ssp_raw_{ano}.xlsx")

    if os.path.exists(destino):
        os.remove(destino)

    url = f"{BASE_URL}/VeiculosSubtraidos_{ano}.xlsx"
    print(f"  ↓ {ano}: baixando de {url}")

    try:
        r = requests.get(url, headers=HEADERS, timeout=300, stream=True)
        if r.status_code == 404:
            print(f"  ✗ {ano}: não disponível (404)")
            return None
        r.raise_for_status()

        total   = int(r.headers.get("content-length", 0))
        baixado = 0
        last_print = 0
        with open(destino, "wb") as f:
            for chunk in r.iter_content(1024 * 1024):
                f.write(chunk)
                baixado += len(chunk)
                # Print a cada 10MB para não poluir os logs
                if baixado - last_print > 10 * 1024 * 1024:
                    if total:
                        print(f"     {baixado/1024/1024:.0f}/{total/1024/1024:.0f} MB")
                    last_print = baixado

        size = os.path.getsize(destino) / 1024 / 1024
        print(f"  ✓ {ano}: {size:.1f} MB baixado")
        time.sleep(2)
        return destino

    except Exception as e:
        print(f"  ✗ {ano}: ERRO — {e}")
        if os.path.exists(destino):
            os.remove(destino)
        return None


def normalizar(path_xlsx, ano):
    """Remove colunas indesejadas, salva XLSX normalizado no destino."""

    wb = openpyxl.load_workbook(path_xlsx, read_only=True, data_only=True)
    aba = next((s for s in wb.sheetnames if "VEICULO" in s.upper()), wb.sheetnames[0])
    wb.close()

    print(f"  📖 {ano}: lendo aba '{aba}'...")
    df = pd.read_excel(path_xlsx, sheet_name=aba, dtype=str, engine="openpyxl")

    df = df[[c for c in df.columns if c is not None and str(c).strip() not in ("None","")]]
    df = df.rename(columns=RENOMEAR)
    df = df.drop(columns=[c for c in df.columns if c in REMOVER], errors="ignore")

    cols_unicas = []
    vistas = set()
    for c in df.columns:
        if c not in vistas:
            cols_unicas.append(c)
            vistas.add(c)
    df = df[cols_unicas]

    os.makedirs(PASTA_DESTINO, exist_ok=True)
    destino = os.path.join(PASTA_DESTINO, f"VeiculosSubtraidos_{ano}.xlsx")

    print(f"  💾 {ano}: salvando {len(df):,} registros × {len(df.columns)} colunas")
    df.to_excel(destino, index=False, engine="openpyxl", sheet_name=f"VEICULOS_{ano}")
    size = os.path.getsize(destino) / 1024 / 1024
    print(f"  ✓ {ano}: salvo ({size:.1f} MB)")

    # Remove arquivo bruto para economizar espaço
    if os.path.exists(path_xlsx):
        os.remove(path_xlsx)

    return destino


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true",
                        help="Força re-download do ano atual")
    args = parser.parse_args()

    print("=" * 55)
    print(f"  SSP-SP — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 55)

    processados = 0
    pulados     = 0

    for ano in ANOS:
        print(f"\n── ANO {ano} ──")

        # Ano atual: sempre re-baixa (tem dados novos todo mês)
        # Ano antigo: só baixa se ainda não existe
        eh_ano_atual = (ano == ANO_ATUAL)
        precisa_baixar = eh_ano_atual or (args.force) or (not ja_existe_normalizado(ano))

        if not precisa_baixar:
            print(f"  ⏭️  {ano} já existe — pulando (dado histórico)")
            pulados += 1
            continue

        path_bruto = baixar(ano)
        if not path_bruto:
            continue

        try:
            normalizar(path_bruto, ano)
            processados += 1
        except Exception as e:
            print(f"  ✗ {ano}: erro — {e}")

    print()
    print("=" * 55)
    print(f"  ✅ {processados} processados · {pulados} já existiam")
    print("=" * 55)


if __name__ == "__main__":
    main()
